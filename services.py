import matplotlib.pyplot as plt
from datetime import datetime, timedelta, date
import database
import os

class AttendanceService:
    def __init__(self):
        database.ensure_data_dir()
        self.reload()

    def reload(self):
        self.users, self.students, self.lecturers = database.load_users()
        self.classes_map = database.load_classes()  # dict class_name -> lecturer
        self.classes = list(self.classes_map.keys())
        self.attendance_records = database.load_attendance_records()  # list of dicts

    # user & class management
    def add_user(self, username, password, role):
        if username in self.users:
            return False, "Username exists"
        database.append_user(username, password, role)
        self.reload()
        return True, "User added"

    def add_class(self, class_name, lecturer_username=""):
        if class_name in self.classes_map:
            return False, "Class exists"
        database.append_class(class_name, lecturer_username)
        self.reload()
        return True, "Class added"
    
    def delete_user(self, username):
        """
        通过 database.delete_user 真正删掉 CSV 里的用户，然后刷新缓存。
        """
        deleted = database.delete_user(username)
        if deleted:
            self.reload()
            return True, f"User '{username}' deleted."
        else:
            return False, f"User '{username}' not found."

    def delete_class(self, class_name):
        """
        通过 database.delete_class 真正删掉 CSV 里的班级，然后刷新缓存。
        """
        deleted = database.delete_class(class_name)
        if deleted:
            self.reload()
            return True, f"Class '{class_name}' deleted."
        else:
            return False, f"Class '{class_name}' not found."


    # attendance
    def mark_attendance(self, class_name, student_username, status="Present"):
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_in = now.strftime("%H:%M:%S")
        # check if already marked for the same date
        for rec in self.attendance_records:
            if rec["date"] == date_str and rec["class_name"] == class_name and rec["student_username"] == student_username:
                return False, "Already marked"
        database.append_attendance(date_str, class_name, student_username, status, time_in)
        self.reload()
        return True, "Marked"

    def update_attendance(self, date_str, class_name, student_username, new_status):
        ok = database.update_attendance_record(date_str, class_name, student_username, new_status)
        if ok:
            self.reload()
            return True
        return False

    # helpers for UI
    def get_attendance_map_for_date(self, target_date=None):
        """Return dict: class_name -> {student_username: status} for target_date (default today)"""
        if target_date is None:
            target_date = datetime.now().strftime("%Y-%m-%d")
        records = [r for r in self.attendance_records if r["date"] == target_date]
        mapping = {}
        for r in records:
            mapping.setdefault(r["class_name"], {})[r["student_username"]] = r["status"]
        return mapping

    def get_class_attendance_stats(self, class_name):
        """Return counts for Present/Absent/Late/Excused across all records for this class."""
        counts = {"Present": 0, "Absent": 0, "Late": 0, "Excused": 0}
        for r in self.attendance_records:
            if r.get("class_name") == class_name:
                status = r.get("status", "Absent")
                counts[status] = counts.get(status, 0) + 1
        return counts

    def plot_attendance_trend(self, class_name):
        """
        Build 14-day attendance rate (Present / total * 100) for the class and save PNG.
        Returns (True, path) or (False, message).
        """
        if not self.attendance_records:
            return False, "No records"

        today = date.today()
        start_date = today - timedelta(days=13)
        # prepare date list
        date_list = [start_date + timedelta(days=i) for i in range(14)]

        # compute totals and presents per date
        totals = {d: 0 for d in date_list}
        presents = {d: 0 for d in date_list}

        for r in self.attendance_records:
            if r.get("class_name") != class_name:
                continue
            try:
                rec_date = datetime.strptime(r.get("date", ""), "%Y-%m-%d").date()
            except Exception:
                continue
            if rec_date < start_date or rec_date > today:
                continue
            totals[rec_date] += 1
            if r.get("status") == "Present":
                presents[rec_date] += 1

        # if all totals zero -> no recent data
        if all(totals[d] == 0 for d in date_list):
            return False, "No recent data"

        # build rates
        rates = []
        x_dates = []
        for d in date_list:
            total = totals[d]
            present = presents[d]
            rate = (present / total * 100) if total > 0 else 0.0
            x_dates.append(d)
            rates.append(rate)

        # plot (matplotlib can plot datetime.date)
        plt.figure(figsize=(8, 4))
        plt.plot(x_dates, rates, marker="o", linestyle="-")
        plt.title(f"14-day Attendance Rate - {class_name}")
        plt.xlabel("Date")
        plt.ylabel("Attendance Rate (%)")
        plt.ylim(0, 100)
        plt.grid(True)
        plt.gcf().autofmt_xdate()
        out = os.path.join(os.path.dirname(__file__), "attendance_trend.png")
        plt.tight_layout()
        plt.savefig(out)
        plt.close()
        return True, out

    def get_student_history(self, student_username):
        """Return list of attendance records for a student sorted by date desc."""
        recs = [r for r in self.attendance_records if r.get("student_username") == student_username]
        # parse dates and sort descending
        def parse_date(rec):
            try:
                return datetime.strptime(rec.get("date", ""), "%Y-%m-%d").date()
            except Exception:
                return date.min
        recs.sort(key=parse_date, reverse=True)
        return recs

    # Excel export helpers using openpyxl (preferred) with fallbacks
    def _write_xlsx_openpyxl(self, rows, headers, out_xlsx, sheet_name="Sheet1"):
        """Helper: write list-of-rows to .xlsx using openpyxl with header formatting."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import NamedStyle

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        header_font = Font(bold=True)
        align = Alignment(horizontal="left", vertical="center")

        # Define a date style for date columns
        date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")

        # write header
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font = header_font
            cell.alignment = align

        # write rows
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = align
                # Apply date style if the header indicates a date column
                if headers[c_idx - 1].lower() == "date" and isinstance(val, (datetime, date)):
                    cell.style = date_style

        # autosize columns
        for i, h in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            max_len = len(str(h))  # Start with header length
            for cell in ws[col_letter]:
                if cell.value is not None:
                    l = len(str(cell.value))
                    if l > max_len:
                        max_len = l
            # Set column width with a small buffer
            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(out_xlsx)
        return out_xlsx

    def export_class_stats_to_excel(self, class_name, out_path=None):
        """
        Export per-date attendance counts for a class to a real .xlsx file.
        Uses openpyxl for formatting. Falls back to CSV if needed.
        """
        # build summary: list of rows with header
        # gather statuses per date
        today = date.today()
        dates = sorted({(datetime.strptime(r["date"], "%Y-%m-%d").date())
                        for r in self.attendance_records if r.get("class_name") == class_name and r.get("date")}, reverse=False)
        if not dates:
            return False, "No records for this class"

        # determine all statuses encountered
        statuses = set()
        for r in self.attendance_records:
            if r.get("class_name") == class_name:
                statuses.add(r.get("status", "Absent"))
        statuses = sorted(statuses)

        # header: date + statuses
        headers = ["date"] + statuses
        rows = []
        for d in dates:
            counts = {s: 0 for s in statuses}
            for r in self.attendance_records:
                if r.get("class_name") != class_name:
                    continue
                try:
                    rec_date = datetime.strptime(r.get("date", ""), "%Y-%m-%d").date()
                except Exception:
                    continue
                if rec_date == d:
                    counts[r.get("status", "Absent")] = counts.get(r.get("status", "Absent"), 0) + 1
            row = [d.strftime("%Y-%m-%d")] + [counts[s] for s in statuses]
            rows.append(row)

        out_dir = os.path.join(os.path.dirname(__file__), "data")
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)

        # use timestamped filename to avoid overwriting files that might be open in Excel
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"{class_name}_attendance_{ts}.xlsx"
        out_xlsx = out_path or os.path.join(out_dir, default_name)

        # try to write .xlsx using openpyxl helper
        try:
            self._write_xlsx_openpyxl(rows, headers, out_xlsx, sheet_name=class_name)
            return True, out_xlsx
        except Exception:
            # fallback to simple CSV
            try:
                # ensure CSV fallback also uses a unique filename
                out_csv = out_xlsx.replace(".xlsx", ".csv")
                with open(out_csv, "w", encoding="utf-8", newline="") as f:
                    f.write(",".join(headers) + "\n")
                    for r in rows:
                        f.write(",".join(str(x) for x in r) + "\n")
                return True, out_csv
            except Exception as e:
                return False, f"Failed to export: {e}"

    def export_student_history_to_excel(self, student_username, out_path=None):
        """Export full attendance history for a student to .xlsx (openpyxl) or CSV fallback."""
        hist = self.get_student_history(student_username)
        if not hist:
            return False, "No records for this student"

        headers = ["date", "class_name", "student_username", "status", "time_in"]
        rows = []
        for r in hist:
            rows.append([r.get("date", ""), r.get("class_name", ""), r.get("student_username", ""), r.get("status", ""), r.get("time_in", "")])

        out_dir = os.path.join(os.path.dirname(__file__), "data")
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"{student_username}_history_{ts}.xlsx"
        out_xlsx = out_path or os.path.join(out_dir, default_name)

        try:
            self._write_xlsx_openpyxl(rows, headers, out_xlsx, sheet_name=(student_username or "History"))
            return True, out_xlsx
        except Exception:
            try:
                out_csv = out_xlsx.replace(".xlsx", ".csv")
                with open(out_csv, "w", encoding="utf-8", newline="") as f:
                    f.write(",".join(headers) + "\n")
                    for r in rows:
                        f.write(",".join(str(x) for x in r) + "\n")
                return True, out_csv
            except Exception as e:
                return False, f"Failed to export: {e}"

    def get_attendance_history_for_class(self, class_name, days=14):
        """
        Return an ordered dict mapping datetime.date -> attendance rate (0-100) for the past `days` days
        (including today). If no data at all, returns empty dict.
        This is intended to be used by the GUI to build animated plots.
        """
        if not self.attendance_records:
            return {}

        today = date.today()
        start_date = today - timedelta(days=days - 1)
        # prepare date list
        date_list = [start_date + timedelta(days=i) for i in range(days)]

        totals = {d: 0 for d in date_list}
        presents = {d: 0 for d in date_list}

        for r in self.attendance_records:
            if r.get("class_name") != class_name:
                continue
            try:
                rec_date = datetime.strptime(r.get("date", ""), "%Y-%m-%d").date()
            except Exception:
                continue
            if rec_date < start_date or rec_date > today:
                continue
            totals[rec_date] += 1
            if r.get("status") == "Present":
                presents[rec_date] += 1

        # if all totals zero -> no recent data
        if all(totals[d] == 0 for d in date_list):
            return {}

        # build rates dict
        rates = {}
        for d in date_list:
            total = totals[d]
            present = presents[d]
            rate = (present / total * 100) if total > 0 else 0.0
            rates[d] = rate

        return rates

    # compatibility aliases (GUI may try different names)
    def get_class_attendance_history(self, class_name, days=14):
        return self.get_attendance_history_for_class(class_name, days=days)

    def get_attendance_history(self, class_name, days=14):
        return self.get_attendance_history_for_class(class_name, days=days)
